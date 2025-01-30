from phe import paillier
import numpy as np
import json
from copy import deepcopy
import os
import random, re
from copy import deepcopy
from os import environ
from time import time
from datetime import datetime
from datetime import timedelta
from collections import defaultdict

import numpy as np

import tenseal as ts
import tenseal.sealapi as sealapi
import openfhe
from openfhe import *
from Pyfhel import Pyfhel
import time
import math
import pandas as pd
from openpyxl import load_workbook

def convert_size(size_bytes):
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])

def append_dict_to_excel(time_df, record_df, avg_df, keysize_df, size_df, avg_size_df, file_name, sheet1='exec_time', sheet2='time_avg', sheet3='op_result', sheet4='keysize', sheet5='ctxtsize', sheet6 = 'size_avg'):
    if not os.path.exists(file_name):
        # If file doesn't exist, create it with the data
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            time_df.to_excel(writer, index=False, sheet_name=sheet1)
            avg_df.to_excel(writer, index=False, sheet_name=sheet2)
            record_df.to_excel(writer, index=False, sheet_name=sheet3)
            keysize_df.to_excel(writer, index=False, sheet_name=sheet4)
            size_df.to_excel(writer, index=False, sheet_name=sheet5)
            avg_size_df.to_excel(writer, index=False, sheet_name=sheet6)
    else:
        # If file exists, append data
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            # Load existing workbook
            writer.book = load_workbook(file_name)
            
            # Append data to sheet1
            if sheet1 in writer.book.sheetnames:
                start_row = writer.book[sheet1].max_row
                time_df.to_excel(writer, index=False, sheet_name=sheet1, header=False, startrow=start_row)
            else:
                time_df.to_excel(writer, index=False, sheet_name=sheet1)
            
            # Append data to sheet2
            if sheet2 in writer.book.sheetnames:
                start_row2 = writer.book[sheet2].max_row
                avg_df.to_excel(writer, index=False, sheet_name=sheet2, header=False, startrow=start_row2)
            else:
                avg_df.to_excel(writer, index=False, sheet_name=sheet2)

            # Append data to sheet3
            if sheet3 in writer.book.sheetnames:
                start_row3 = writer.book[sheet3].max_row
                record_df.to_excel(writer, index=False, sheet_name=sheet3, header=False, startrow=start_row3)
            else:
                record_df.to_excel(writer, index=False, sheet_name=sheet3)

            # Append data to sheet4
            if sheet4 in writer.book.sheetnames:
                start_row4 = writer.book[sheet4].max_row
                keysize_df.to_excel(writer, index=False, sheet_name=sheet4, header=False, startrow=start_row4)
            else:
                keysize_df.to_excel(writer, index=False, sheet_name=sheet4)

            # Append data to sheet5
            if sheet5 in writer.book.sheetnames:
                start_row5 = writer.book[sheet5].max_row
                size_df.to_excel(writer, index=False, sheet_name=sheet5, header=False, startrow=start_row5)
            else:
                size_df.to_excel(writer, index=False, sheet_name=sheet5)

            # Append data to sheet6
            if sheet6 in writer.book.sheetnames:
                start_row6 = writer.book[sheet6].max_row
                avg_size_df.to_excel(writer, index=False, sheet_name=sheet6, header=False, startrow=start_row6)
            else:
                avg_size_df.to_excel(writer, index=False, sheet_name=sheet6)             

                   

# for measuring difference of message and decrypted plaintext in CKKS scheme 
# Source: https://github.com/openfheorg/openfhe-python/blob/main/examples/pke/iterative-ckks-bootstrapping.py                        
def calculate_approximation_error(result, expected_result):
    # result, plaintext after decryption
    # expected_result, message results

    if len(result) != len(expected_result):
        raise Exception("Cannot compare vectors with different numbers of elements")
    # using the infinity norm
    # error is abs of the difference of real parts
    diff_arr = [abs(el1.real - el2.real) for (el1, el2) in zip(result, expected_result)]
    max_error = max(diff_arr)
    max_error_idx = np.argmax(diff_arr)
    min_error = min(diff_arr)
    min_error_idx = np.argmin(diff_arr)
    # return absolute value of log base2 of the error

    # TODO: what does this mean in math (numbers)
    # print(f"Bootstrapping precision after 1 iteration: {abs(math.log(max_error, 2))} bits\n")

    # Measure the overall difference in Mean Squared Error (MSE)
    diff_arr_MSE = [abs(el1.real - el2.real)**2 for (el1, el2) in zip(result, expected_result)]
    MSE = sum(diff_arr_MSE)/ len(result)

    if MSE != 0:
        return max_error, min_error, MSE, abs(math.log(max_error,2))
    else:
        return max_error, min_error, MSE, 0 # avoid math domain error
    
# def key_gen_Paillier():
#         # TODO: parameterize or put in file
#         Pyhon_Paillier_config = {
#             'n_clients': 1,
#             'key_length': 256, # 64, 128: overflow # 256
#             'n_iter': 50,
#             'eta': 1.5,
#         }
        
#         keypair = paillier.generate_paillier_keypair(n_length=Pyhon_Paillier_config['key_length'])
#         pubKey, privKey = keypair

#         return pubKey, privKey

def key_gen_OpenFHE_CKKS(ring_dim, scale_bit, datafolder):
    # print("key gen in OpenFHE_CKKS")
    keysize_dict = {}
    # For benchmarking: set security level as 128-bit, max size of q as 218, and N (degree of polynomial; ring dim) as 8192
    # https://github.com/openfheorg/openfhe-python/blob/main/examples/pke/simple-real-numbers.py
    # https://github.com/openfheorg/openfhe-python/blob/main/examples/pke/advanced-real-numbers.py

    # set as 128-bit security level
    # https://openfhe.discourse.group/t/why-do-my-parameters-not-comply-with-128-bit-security/432
    
    # Q0 + mult_depth* Qi should equal to maximum size of Q in the table

    mult_depth = 1
    first_mod_size = 60 # Q0; first_mod_size can't exceed 60
    scale_mod_size = scale_bit # assigned by options.py


    # if ring_dim == 8192: # logQ = 218
    #     # for ring_dim = 8192; Q = 218: 60 + 49*2 + 60 = 218
    #     # mult_depth = 2
    #     # scale_mod_size = 49 # Qi
    #     # first_mod_size = 60 # Q0

    #     scale_mod_size = 52 # Qi
    #     first_mod_size = 60 # Q0; first_mod_size can't exceed 60


    num_data_points = (int)(ring_dim/2)
    batch_size_power = math.ceil(math.log(num_data_points, 2)) 
    batch_size = 2** batch_size_power # find the most appropriate batchsize for arrays in experiment (power of 2)
    # * batchsize < SetRingDim, also batchsize > #params in the layer

    parameters = CCParamsCKKSRNS()
    # parameters.SetSecurityLevel(SecurityLevel.HEStd_128_classic)
    parameters.SetSecurityLevel(SecurityLevel.HEStd_NotSet) # some parameter combinations would not satisfy HE standard, use NotSet to 
    parameters.SetMultiplicativeDepth(mult_depth)
    parameters.SetScalingModSize(scale_mod_size) # Qi
    parameters.SetFirstModSize(first_mod_size) # Q0
    parameters.SetRingDim(ring_dim)    
    parameters.SetBatchSize(batch_size)
    parameters.SetKeySwitchTechnique(KeySwitchTechnique.BV) # make the ring dim fits HE standards
    
    cc = GenCryptoContext(parameters)
    
    cc.Enable(PKESchemeFeature.PKE)
    cc.Enable(PKESchemeFeature.KEYSWITCH) # relinearlization, used in TenSeal
    cc.Enable(PKESchemeFeature.LEVELEDSHE)

    
    # print("The CKKS scheme is using ring dimension: " + str(cc.GetRingDimension()))
    # print("The CKKS scheme is using mudulus (q): " + str(cc.GetModulus()))
    # print("The CKKS scheme is using mudulus (log_2(q)): " + str(math.log(cc.GetModulus(), 2)))
    # print("The CKKS scheme is using first mudulus (q0, getModulusCKKS): " + str(cc.GetModulusCKKS()))
    # print("The CKKS scheme is using first mudulus (q0, GetFirstModSize): " + str(parameters.GetFirstModSize()))
    # print("The CKKS scheme is using first mudulus (log_2(q0), getModulusCKKS): " + str(math.log(cc.GetModulusCKKS(), 2)))
    # print("The CKKS scheme is using scale bit (q1, GetScalingModSize): " + str(parameters.GetScalingModSize()))
    # print("The CKKS scheme is using GetPlaintextModulus: " + str(cc.GetPlaintextModulus()))

    # print("The CKKS scheme is using GetMultiplicativeDepth: " + str(parameters.GetMultiplicativeDepth()))
    # print("The CKKS scheme is using security level: " + str(parameters.GetSecurityLevel()))
    # print("The CKKS scheme is using KeySwitchTechnique: " + str(parameters.GetKeySwitchTechnique()))
    # print("The CKKS scheme is using BatchSize: " + str(parameters.GetBatchSize()))
    
    keygen_t_start = time.time()
    keys = cc.KeyGen()
    cc.EvalMultKeyGen(keys.secretKey) # required if needing multiplying ciphertexts, EVK
    pubKey, privKey = keys.publicKey, keys.secretKey
    keygen_t_end = time.time()
    keygen_time = keygen_t_end - keygen_t_start
    # print(f"Time taken for keygen (second): {keygen_time}")

    if not SerializeToFile(datafolder + "/CKKS_context.txt", cc, BINARY):
        raise Exception("Error writing serialization of CKKS_context to CKKS_context.txt")
    print("CKKS_context has been serialized.")
    CKKS_context_stats = os.stat(datafolder + "/CKKS_context.txt")
    print(f"CKKS_context size: {CKKS_context_stats.st_size} Bytes")
    keysize_dict["context_size"] = CKKS_context_stats.st_size

    if not SerializeToFile(datafolder + "/pubKey.txt", pubKey, BINARY):
        raise Exception("Error writing serialization of pubKey to pubKey.txt")
    print("pubKey has been serialized.")
    pubKey_stats = os.stat(datafolder + "/pubKey.txt")
    print(f"pubKey size: {pubKey_stats.st_size} Bytes")
    keysize_dict["pubKey_size"] = pubKey_stats.st_size

    if not SerializeToFile(datafolder + "/privKey.txt", privKey, BINARY):
        raise Exception("Error writing serialization of privKey to privKey.txt")
    print("privKey has been serialized.")
    privKey_stats = os.stat(datafolder + "/privKey.txt")
    print(f"privKey size: {privKey_stats.st_size} Bytes")
    keysize_dict["privKey_size"] = privKey_stats.st_size

    return pubKey, privKey, cc, keygen_time, keysize_dict

# TenSeal
def key_gen_TenSeal_CKKS(ring_dim, scale_bit, datafolder):
    keysize_dict = {}

    coeff_mod_bit_sizes_arr = [60, scale_bit, 60] # 60* 2 + scale_bit

    # if ring_dim == 8192:
    #     # coeff_mod_bit_sizes_arr = [60, 49, 49, 60] # 60* 2 + 49* 2
    #     # bits_scale = 49
    #     coeff_mod_bit_sizes_arr = [60, 52, 60] # 60* 2 + 49* 2
    #     bits_scale = 52

   
    context = ts.context(
        ts.SCHEME_TYPE.CKKS, 
        poly_modulus_degree=ring_dim,
        coeff_mod_bit_sizes= coeff_mod_bit_sizes_arr
    )
    context.global_scale = pow(2, scale_bit)

    context.security_level = 0

    keygen_t_start = time.time()
    context.generate_galois_keys() # https://dp-ml.github.io/2021-workshop-ICLR/files/18.pdf galois: for rotation
    context.generate_relin_keys()
    pubKey = context.public_key()
    privKey = context.secret_key()
    keygen_t_end = time.time()
    keygen_time = keygen_t_end - keygen_t_start
    # print(f"Time taken for keygen (second): {keygen_time}")

    # https://github.com/serengil/tensorflow-101/blob/master/python/Homomorphic-Face-Recognition.ipynb
    ser_context = context.serialize(save_public_key=True, save_secret_key=True, )
    with open(os.path.join(datafolder, "ser_context"), "wb") as f:
        f.write(ser_context)
    print("ser_context has been serialized.")
    CKKS_context_stats = os.stat(os.path.join(datafolder, "ser_context"))
    print(f"CKKS_context size: {CKKS_context_stats.st_size} Bytes")
    keysize_dict["context_size"] = CKKS_context_stats.st_size    

    ser_privKey = context.serialize(save_secret_key=True)
    with open(os.path.join(datafolder, "ser_privKey"), "wb") as f:
        f.write(ser_privKey)
    print("ser_privKey has been serialized.")
    privKey_stats = os.stat(os.path.join(datafolder, "ser_privKey"))
    print(f"privKey size: {privKey_stats.st_size} Bytes")
    keysize_dict["privKey_size"] = privKey_stats.st_size

    # context.make_context_public() #drop the secret_key from the context
    ser_pubKey = context.serialize(save_public_key=True)
    with open(os.path.join(datafolder, "ser_pubKey"), "wb") as f:
        f.write(ser_pubKey)
    print("ser_pubKey has been serialized.")
    pubKey_stats = os.stat(os.path.join(datafolder, "ser_pubKey"))
    print(f"pubKey size: {pubKey_stats.st_size} Bytes")
    keysize_dict["pubKey_size"] = pubKey_stats.st_size
    
    return pubKey, privKey, context, keygen_time, keysize_dict

# def key_gen_TenSeal_CKKS(ring_dim, scale_bit, datafolder):
#     keysize_dict = {}

#     coeff_mod_bit_sizes_arr = [60, scale_bit, 60] # 60* 2 + scale_bit

#     # if ring_dim == 8192:
#     #     # coeff_mod_bit_sizes_arr = [60, 49, 49, 60] # 60* 2 + 49* 2
#     #     # bits_scale = 49
#     #     coeff_mod_bit_sizes_arr = [60, 52, 60] # 60* 2 + 49* 2
#     #     bits_scale = 52

#     parms = sealapi.EncryptionParameters(sealapi.SCHEME_TYPE.CKKS)
#     parms.set_poly_modulus_degree(ring_dim)
#     coeff = sealapi.CoeffModulus.Create(ring_dim, coeff_mod_bit_sizes_arr)
#     parms.set_coeff_modulus(coeff)
#     context = sealapi.SEALContext(parms, False, sealapi.SEC_LEVEL_TYPE.NONE)
   
#     # context = ts.context(
#     #     ts.SCHEME_TYPE.CKKS, 
#     #     poly_modulus_degree=ring_dim,
#     #     coeff_mod_bit_sizes= coeff_mod_bit_sizes_arr
#     #     # SEC_LEVEL_TYPE=sealapi.SEC_LEVEL_TYPE.TC128 # allow unreasonable ring dimension and Qi that are not suitable for HE standard
#     # )
#     # context.global_scale = pow(2, scale_bit)

#     keygen_t_start = time.time()
#     # context.generate_galois_keys() # https://dp-ml.github.io/2021-workshop-ICLR/files/18.pdf galois: for rotation
#     keygen = sealapi.KeyGenerator(context)
#     # galois_keys = sealapi.GaloisKeys()
#     # context.generate_relin_keys()
#     # pubKey = context.public_key()
#     pubKey = sealapi.PublicKey()
#     keygen.create_public_key(pubKey)
#     privKey = keygen.secret_key()
#     # privKey = context.secret_key()
#     keygen_t_end = time.time()
#     keygen_time = keygen_t_end - keygen_t_start
#     # print(f"Time taken for keygen (second): {keygen_time}")

#     # # https://github.com/serengil/tensorflow-101/blob/master/python/Homomorphic-Face-Recognition.ipynb
#     # ser_context = context.serialize(save_public_key=True, save_secret_key=True, )
#     # with open(os.path.join(datafolder, "ser_context"), "wb") as f:
#     #     f.write(ser_context)
#     # print("ser_context has been serialized.")
#     # CKKS_context_stats = os.stat(os.path.join(datafolder, "ser_context"))
#     # print(f"CKKS_context size: {CKKS_context_stats.st_size} Bytes")
#     # keysize_dict["context_size"] = CKKS_context_stats.st_size    

#     # ser_privKey = context.serialize(save_secret_key=True)
#     # with open(os.path.join(datafolder, "ser_privKey"), "wb") as f:
#     #     f.write(ser_privKey)
#     # print("ser_privKey has been serialized.")
#     # privKey_stats = os.stat(os.path.join(datafolder, "ser_privKey"))
#     # print(f"privKey size: {privKey_stats.st_size} Bytes")
#     # keysize_dict["privKey_size"] = privKey_stats.st_size

#     # # context.make_context_public() #drop the secret_key from the context
#     # ser_pubKey = context.serialize(save_public_key=True)
#     # with open(os.path.join(datafolder, "ser_pubKey"), "wb") as f:
#     #     f.write(ser_pubKey)
#     # print("ser_pubKey has been serialized.")
#     # pubKey_stats = os.stat(os.path.join(datafolder, "ser_pubKey"))
#     # print(f"pubKey size: {pubKey_stats.st_size} Bytes")
#     # keysize_dict["pubKey_size"] = pubKey_stats.st_size
    
#     return pubKey, privKey, context, keygen_time, keysize_dict

# Pyfhel
def key_gen_Pyfhel_CKKS(ring_dim, scale_bit, datafolder):
    keysize_dict = {}
    Pyfhel_HE = Pyfhel()  # Creating empty Pyfhel object
        
    coeff_mod_bit_sizes_arr = [60, scale_bit, 60] # 60* 2 + scale_bit

    # if ring_dim == 8192:
    #     # bits_scale = 49
    #     # coeff_mod_bit_sizes_arr = [60, 49, 49, 60] # 60*2 + 49* 2 = 218
    #     bits_scale = 52
    #     coeff_mod_bit_sizes_arr = [60, 52, 60] # 60*2 + 52 = 172, mimic FedML-HE


    ckks_params = {
        'scheme': 'CKKS',   # can also be 'ckks'
        'n': ring_dim,         # Polynomial modulus degree. For CKKS, n/2 values can be
                            #  encoded in a single ciphertext.
                            #  Typ. 2^D for D in [10, 15]
        'scale': 2**scale_bit,     # All the encodings will use it for float->fixed point
                            #  conversion: x_fix = round(x_float * scale)
                            #  You can use this as default scale or use a different
                            #  scale on each operation (set in Pyfhel_HE.encryptFrac)
        'qi_sizes': coeff_mod_bit_sizes_arr, # Number of bits of each prime in the chain.
                            # Intermediate values should be  close to log2(scale)
                            # for each operation, to have small rounding errors.
        'sec': 0 # Security of the context parameters (bits). 0 -> None
        # src: https://github.com/ibarrond/Pyfhel/blob/14c35f975fa1ed0a97685535582c10beb77145eb/Pyfhel/Afhel/Afseal.h#L86
        # src: https://github.com/ibarrond/Pyfhel/blob/14c35f975fa1ed0a97685535582c10beb77145eb/Pyfhel/Afhel/Afseal.cpp#L75
                            
    }
    Pyfhel_HE.contextGen(**ckks_params)  # Generate context for ckks scheme

    keygen_t_start = time.time()
    Pyfhel_HE.keyGen()             # Key Generation: generates a pair of public/secret keys
    # https://github.com/ibarrond/Pyfhel/blob/master/examples/Demo_5_CS_Client.py
    pubKey = Pyfhel_HE.to_bytes_public_key()
    privKey = Pyfhel_HE.to_bytes_secret_key()
    context = Pyfhel_HE.to_bytes_context()
    # relinearKey = Pyfhel_HE.to_bytes_relin_key()
    keygen_t_end = time.time()
    keygen_time = keygen_t_end - keygen_t_start


    with open(os.path.join(datafolder, "pubKey"), "wb") as f:
        f.write(pubKey)
    print("pubKey has been serialized.")

    with open(os.path.join(datafolder, "privKey"), "wb") as f:
        f.write(privKey)
    print("privKey has been serialized.")

    with open(os.path.join(datafolder, "context"), "wb") as f:
        f.write(context)
    print("context has been serialized.")

    CKKS_pubKey_stats = os.stat(os.path.join(datafolder, "pubKey"))
    print(f"pubKey_size size: {CKKS_pubKey_stats.st_size} Bytes")
    keysize_dict["pubKey_size"] = CKKS_pubKey_stats.st_size

    CKKS_privKey_stats = os.stat(os.path.join(datafolder, "privKey"))
    print(f"privKey_size size: {CKKS_privKey_stats.st_size} Bytes")
    keysize_dict["privKey_size"] = CKKS_privKey_stats.st_size

    CKKS_context_stats = os.stat(os.path.join(datafolder, "context"))
    print(f"CKKS_context size: {CKKS_context_stats.st_size} Bytes")
    keysize_dict["context_size"] = CKKS_context_stats.st_size

    pkContext = Pyfhel()
    skContext = Pyfhel()
    
    with open(os.path.join(datafolder, "context"), "rb") as f:
        contextFile = f.read()
    
    with open(os.path.join(datafolder, "pubKey"), "rb") as f:
        pubFile = f.read()
    with open(os.path.join(datafolder, "privKey"), "rb") as f:
        privFile = f.read()

    pkContext.from_bytes_context(contextFile)
    pkContext.from_bytes_public_key(pubFile)

    skContext.from_bytes_context(contextFile)
    skContext.from_bytes_public_key(pubFile)
    skContext.from_bytes_secret_key(privFile)
    
    # pkContext.qi = Pyfhel_HE.get_qi() # attribute 'qi' of 'Pyfhel.Pyfhel.Pyfhel' objects is not writable
    # skContext.qi = Pyfhel_HE.get_qi()

    return pkContext, skContext, Pyfhel_HE, keygen_time, keysize_dict

# # Palisade
# def key_gen_Palisade_CKKS(HE_method, num_data_points, ring_dim, datafolder):
#     keysize_dict = {}
#     if ring_dim == 8192:
#         # for ring_dim = 8192; Q = 218
#         mult_depth = 2
#         scale_mod_size = 49 # 52; Qi
#         first_mod_size = 60 # Q0

#     # elif ring_dim == 16384:
#     #     # for ring_dim = 16384; Q = 438
#     #     mult_depth = 8
#     #     scale_mod_size = 40 # 52; Qi
#     #     first_mod_size = 59 # Q0 # because first_mod_size can't exceed 60, increase multiplication depth

#     # elif ring_dim == 32768:
#     #     # for ring_dim = 32768; Q = 881
#     #     mult_depth = 16
#     #     scale_mod_size = 47 # 52; Qi
#     #     first_mod_size = 59 # Q0 # because first_mod_size can't exceed 60, increase multiplication depth

#     batch_size = (int)(ring_dim/2)
#     keygen_t_start = time.time()
#     crypto = pycrypto.CKKSwrapper()
#     crypto.KeyGen(mult_depth, scale_mod_size, batch_size, ring_dim)
#     keygen_t_end = time.time()
#     keygen_time = keygen_t_end - keygen_t_start
#     # print(f"Time taken for keygen (second): {keygen_time}")
#     # crypto.KeyGen(mult_depth, scale_mod_size, 8192, 16384)
#     # crypto.KeyGen(1, 40, 10, 16384)
#     # print(f"crypto.getRingDim(): {crypto.GetRingDim()}; type: {type(crypto.GetRingDim())}")

#     # No method to write context as byte file
#     # with open(os.path.join(datafolder, "context"), "wb") as f:
#     #     f.write(crypto)
#     # print("context has been serialized.")

#     # keysize_dict["pubKey_size"] = "NA"
#     # keysize_dict["privKey_size"] = "NA"
#     # CKKS_context_stats = os.stat(os.path.join(datafolder, "context"))
#     # print(f"CKKS_context size: {CKKS_context_stats.st_size} Bytes")
#     # keysize_dict["CKKS_context_size"] = CKKS_context_stats.st_size

#     keysize_dict["context_size"] = -1
#     keysize_dict["pubKey_size"] = -1
#     keysize_dict["privKey_size"] = -1
#     keygen_time = -1

#     return crypto, keygen_time, keysize_dict